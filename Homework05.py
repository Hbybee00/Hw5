#Name : Hayden Bybee
#Lab Section 16
#Date : 11/18/24
#Sources:
#N/A


import openpyxl
from openpyxl.styles import Color, PatternFill
import string
wb = openpyxl.Workbook()
sheet = wb.active
colors = {"outlinepink" : Color(rgb ="f0a8df"), "shadingpink" : Color(rgb ="ebb9df"),
"bodypink" : Color(rgb ="fad2f1"), "eye1_mouthblack" : Color(rgb ="393939"), "eye2" : Color(rgb ="342d47"),
"eye3" : Color(rgb ="34206b"), "feet" : Color(rgb ="e85d8e"),
"tongue" : Color(rgb ="854545")}

backgroundblue = "a8bfe3"

colorplacements = {"outlinepink" : ['A11','A12','B7','B8','B9','B10','B13','B14','B15','B16','C6','C17','D5','D18','E5','E18','F6','F17','G7','G8','G9','G10','G13','G14','G15','G16','H11','H12'],
                   "bodypink":['B11','C7','C11','C12','C13','C14','C15','D6','D7','D8','D9','D10','D11','D15','D16','E6','E7','E8','E9','E10','E11','E15','E16','F7','F11','F12','F13','G11'],
                   "shadingpink":['B12','C16','D17','E17','F14','F15','F16','G12'], "eye1_mouthblack":['C8','D12','D13','E12','E13','F8'], "eye2":['C9','F9'],"eye3":['C10','F10'],
                   "feet":['B18','B19','C18','C19','F18','F19','G18','G19'], "tongue":['D14','E14']}

fill2 = PatternFill(patternType='solid',fgColor=backgroundblue)

for chr in string.ascii_uppercase[:8]:
    sheet.column_dimensions[chr].width = 8
    for i in range(1,23):
        sheet.row_dimensions[i].height= 20
        coord = chr+str(i)
        cell_filled = False
        for color, code in colors.items():
            fill = PatternFill(patternType='solid',fgColor=code)
            if coord in colorplacements.get(color,[]):
                sheet[coord].fill = fill
                cell_filled = True
            if not cell_filled:
                sheet[coord].fill = fill2

wb.save('drawing.xlsx')